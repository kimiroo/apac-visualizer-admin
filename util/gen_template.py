from typing import TYPE_CHECKING

import os
import base64

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side

from const.file_path import *

if TYPE_CHECKING:
    from openpyxl.workbook.workbook import _WorksheetOrChartsheetLike

def gen_template(config: dict) -> str:

    temp_path = '/tmp/template.xlsx' if IS_CONTAINER else BASE_PATH / 'template.xlsx'

    # Workbook
    doc = Workbook()

    # Style
    alignment_center = Alignment(horizontal='center', vertical='center')
    side = Side(border_style='double')
    border = Border()

    ##############
    ### Region ###
    ##############

    sheet: _WorksheetOrChartsheetLike = doc.active
    sheet.title = config['source']['sheet']['region']['name']
    sheet.freeze_panes = 'D3'

    # Country
    sheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    sheet.cell(row=1, column=1).value = 'Country'
    sheet.cell(row=3, column=1).value = 'String'
    sheet.cell(row=1, column=1).alignment = alignment_center

    # Region
    sheet.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)
    sheet.cell(row=1, column=2).value = 'Region'
    sheet.cell(row=3, column=2).value = 'String'
    sheet.cell(row=1, column=2).alignment = alignment_center

    # Remarks
    sheet.merge_cells(start_row=1, start_column=3, end_row=2, end_column=3)
    sheet.cell(row=1, column=3).value = 'Remarks'
    sheet.cell(row=3, column=3).value = 'String'
    sheet.cell(row=1, column=3).alignment = alignment_center

    # Verticals
    for idx, vertical in enumerate(config['vertical'] + ['Total']):
        base_idx = 4 + (idx * 4)

        # Header
        sheet.merge_cells(start_row=1, start_column=base_idx, end_row=1, end_column=(base_idx + 3))
        sheet.cell(row=1, column=base_idx).value = vertical
        sheet.cell(row=1, column=base_idx).alignment = alignment_center

        # Project Count
        sheet.cell(row=2, column=(base_idx + 0)).value = 'Project Count'
        sheet.cell(row=3, column=(base_idx + 0)).value = 'Integer'
        sheet.cell(row=2, column=(base_idx + 0)).alignment = alignment_center

        # Projected Dealer Revenue
        sheet.cell(row=2, column=(base_idx + 1)).value = 'Projected Dealer Revenue'
        sheet.cell(row=3, column=(base_idx + 1)).value = 'Decimal'
        sheet.cell(row=2, column=(base_idx + 1)).alignment = alignment_center

        # Potential Market Value
        sheet.cell(row=2, column=(base_idx + 2)).value = 'Potential Market Value'
        sheet.cell(row=3, column=(base_idx + 2)).value = 'Decimal'
        sheet.cell(row=2, column=(base_idx + 2)).alignment = alignment_center

        # Total Project Value
        sheet.cell(row=2, column=(base_idx + 3)).value = 'Total Project Value'
        sheet.cell(row=3, column=(base_idx + 3)).value = 'Decimal'
        sheet.cell(row=2, column=(base_idx + 3)).alignment = alignment_center

    # Apply border
    for column_idx in range(sheet.max_column):
        sheet.cell(row=2, column=(column_idx + 1)).border = border

    ##############
    ### Dealer ###
    ##############

    sheet: _WorksheetOrChartsheetLike = doc.create_sheet(config['source']['sheet']['dealer']['name'])
    sheet.freeze_panes = 'A2'

    # Fixed columns
    sheet.cell(row=1, column=1).value = 'Area'
    sheet.cell(row=2, column=1).value = 'String'

    sheet.cell(row=1, column=2).value = 'Country'
    sheet.cell(row=2, column=2).value = 'String'

    sheet.cell(row=1, column=3).value = 'Sales Organization'
    sheet.cell(row=2, column=3).value = 'String'

    sheet.cell(row=1, column=4).value = 'ID'
    sheet.cell(row=2, column=4).value = 'String'

    sheet.cell(row=1, column=5).value = 'Name'
    sheet.cell(row=2, column=5).value = 'String'

    sheet.cell(row=1, column=6).value = 'Tier'
    sheet.cell(row=2, column=6).value = 'String'

    sheet.cell(row=1, column=7).value = 'Profile'
    sheet.cell(row=2, column=7).value = 'String'

    sheet.cell(row=1, column=8).value = 'Remarks'
    sheet.cell(row=2, column=8).value = 'String'

    sheet.cell(row=1, column=9).value = 'Address'
    sheet.cell(row=2, column=9).value = 'String'

    sheet.cell(row=1, column=10).value = 'Latitude'
    sheet.cell(row=2, column=10).value = 'Decimal'

    sheet.cell(row=1, column=11).value = 'Longtitude'
    sheet.cell(row=2, column=11).value = 'Decimal'

    sheet.cell(row=1, column=12).value = 'Projected Revenue'
    sheet.cell(row=2, column=12).value = 'Decimal'

    sheet.cell(row=1, column=13).value = 'Actual Revenue'
    sheet.cell(row=2, column=13).value = 'Decimal'

    # Verticals
    for idx, vertical in enumerate(config['vertical']):
        base_idx = 14 + idx
        sheet.cell(row=1, column=base_idx).value = vertical
        sheet.cell(row=2, column=base_idx).value = 'Boolean'

    # Apply Styles
    for column_idx in range(sheet.max_column):
        sheet.cell(row=1, column=(column_idx + 1)).alignment = alignment_center
        sheet.cell(row=1, column=(column_idx + 1)).border = border

    #######################
    ### Dealer Customer ###
    #######################

    sheet: _WorksheetOrChartsheetLike = doc.create_sheet(config['source']['sheet']['dealerCustomer']['name'])
    sheet.freeze_panes = 'A2'

    # Fixed columns
    sheet.cell(row=1, column=1).value = 'Dealer ID'
    sheet.cell(row=2, column=1).value = 'String'

    sheet.cell(row=1, column=2).value = 'Dealer Name'
    sheet.cell(row=2, column=2).value = 'String'

    sheet.cell(row=1, column=3).value = 'Name'
    sheet.cell(row=2, column=3).value = 'String'

    sheet.cell(row=1, column=4).value = 'Sale Value'
    sheet.cell(row=2, column=4).value = 'Decimal'

    sheet.cell(row=1, column=5).value = 'Sale Date'
    sheet.cell(row=2, column=5).value = 'Date'

    sheet.cell(row=1, column=6).value = 'Sale Model'
    sheet.cell(row=2, column=6).value = 'String'

    # Apply Styles
    for column_idx in range(sheet.max_column):
        sheet.cell(row=1, column=(column_idx + 1)).alignment = alignment_center
        sheet.cell(row=2, column=(column_idx + 1)).border = border

    ###################
    ### Key Account ###
    ###################

    sheet: _WorksheetOrChartsheetLike = doc.create_sheet(config['source']['sheet']['keyAccount']['name'])
    sheet.freeze_panes = 'A2'

    # Fixed columns
    sheet.cell(row=1, column=1).value = 'ID'
    sheet.cell(row=2, column=1).value = 'String'

    sheet.cell(row=1, column=2).value = 'Account Name'
    sheet.cell(row=2, column=2).value = 'String'

    sheet.cell(row=1, column=3).value = 'Name'
    sheet.cell(row=2, column=3).value = 'String'

    sheet.cell(row=1, column=4).value = 'Address'
    sheet.cell(row=2, column=4).value = 'String'

    sheet.cell(row=1, column=5).value = 'Remarks'
    sheet.cell(row=2, column=5).value = 'String'

    sheet.cell(row=1, column=6).value = 'Latitude'
    sheet.cell(row=2, column=6).value = 'Decimal'

    sheet.cell(row=1, column=7).value = 'Longtitude'
    sheet.cell(row=2, column=7).value = 'Decimal'

    sheet.cell(row=1, column=8).value = 'Value'
    sheet.cell(row=2, column=8).value = 'Decimal'

    sheet.cell(row=1, column=9).value = 'Water Consumption'
    sheet.cell(row=2, column=9).value = 'Decimal'

    sheet.cell(row=1, column=10).value = 'Is Customer'
    sheet.cell(row=2, column=10).value = 'Boolean'

    # Verticals
    for idx, vertical in enumerate(config['vertical']):
        base_idx = 11 + idx
        sheet.cell(row=1, column=base_idx).value = vertical
        sheet.cell(row=2, column=base_idx).value = 'Boolean'

    # Apply Styles
    for column_idx in range(sheet.max_column):
        sheet.cell(row=1, column=(column_idx + 1)).alignment = alignment_center
        sheet.cell(row=1, column=(column_idx + 1)).border = border

    #######################
    ### Priority Target ###
    #######################

    sheet: _WorksheetOrChartsheetLike = doc.create_sheet(config['source']['sheet']['priorityTarget']['name'])
    sheet.freeze_panes = 'A2'

    # Fixed columns
    sheet.cell(row=1, column=1).value = 'Area'
    sheet.cell(row=2, column=1).value = 'String'

    sheet.cell(row=1, column=2).value = 'Country'
    sheet.cell(row=2, column=2).value = 'String'

    sheet.cell(row=1, column=3).value = 'Region'
    sheet.cell(row=2, column=3).value = 'String'

    sheet.cell(row=1, column=4).value = 'ID'
    sheet.cell(row=2, column=4).value = 'String'

    sheet.cell(row=1, column=5).value = 'Name'
    sheet.cell(row=2, column=5).value = 'String'

    sheet.cell(row=1, column=6).value = 'Address'
    sheet.cell(row=2, column=6).value = 'String'

    sheet.cell(row=1, column=7).value = 'Remarks'
    sheet.cell(row=2, column=7).value = 'String'

    sheet.cell(row=1, column=8).value = 'Latitude'
    sheet.cell(row=2, column=8).value = 'Decimal'

    sheet.cell(row=1, column=9).value = 'Longtitude'
    sheet.cell(row=2, column=9).value = 'Decimal'

    sheet.cell(row=1, column=10).value = 'Value'
    sheet.cell(row=2, column=10).value = 'Decimal'

    sheet.cell(row=1, column=11).value = 'Water Consumption'
    sheet.cell(row=2, column=11).value = 'Decimal'

    sheet.cell(row=1, column=12).value = 'Is Customer'
    sheet.cell(row=2, column=12).value = 'Boolean'

    # Verticals
    for idx, vertical in enumerate(config['vertical']):
        base_idx = 13 + idx
        sheet.cell(row=1, column=base_idx).value = vertical
        sheet.cell(row=2, column=base_idx).value = 'Boolean'

    # Apply Styles
    for column_idx in range(sheet.max_column):
        sheet.cell(row=1, column=(column_idx + 1)).alignment = alignment_center
        sheet.cell(row=1, column=(column_idx + 1)).border = border

    #########################
    ### Convert to BASE64 ###
    #########################

    doc.save(temp_path)

    with open(temp_path, 'rb') as f:
        b64 = base64.b64encode(f.read()).decode()

    os.remove(temp_path)

    return b64