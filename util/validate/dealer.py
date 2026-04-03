from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from openpyxl.workbook.workbook import _WorksheetOrChartsheetLike

def check_dealer_sheet(sheet: _WorksheetOrChartsheetLike, config: dict) -> list:
    error_list = [] # {'level': level, 'sheet': sheet, 'cell': cell, 'message': message}

    expected_max_col_cnt = 13 + len(config['vertical'])

    ### Column count check
    if sheet.max_column != expected_max_col_cnt:
        error_list.append({
            'level': 'WARNING',
            'sheet': sheet.title,
            'cell': '-',
            'message': f'Column count mismatch. (Found: {sheet.max_column}, Expected: {expected_max_col_cnt})'
        })

    header_list = [
        'Area', 'Country', 'Sales Organization',
        'ID', 'Name', 'Tier', 'Profile', 'Remarks',
        'Address', 'Latitude', 'Longtitude', 'Projected Revenue',
        'Actual Revenue'
    ]

    ### Header data check
    for idx, header in enumerate(header_list):
        col_idx = 1 + idx
        cell = sheet.cell(row=1, column=col_idx)

        if cell.value != header:
            error_list.append({
                'level': 'WARNING',
                'sheet': sheet.title,
                'cell': cell.coordinate,
                'message': f'Unexpected header value. Subsequent data may be invalid. (Found: {f'"{cell.value}"' if cell.value else 'None'}, Expected: "{header}")'
            })

    ### Vertical header data check
    headers = headers = {cell.value: i for i, cell in enumerate(sheet[1])}
    vertical_list = config['vertical']

    for vertical in vertical_list:
        header_idx = headers.get(vertical, None)

        if header_idx is None:
            error_list.append({
                'level': 'CRITICAL',
                'sheet': sheet.title,
                'cell': '(Header Row)',
                'message': f'Required column "{vertical}" could not be found in the sheet headers.'
            })

    ### Data check
    tier_list = [obj['name'] for obj in config['tiers']]

    for row in sheet.iter_rows(3):

        if row[3].value is None:
            error_list.append({
                'level': 'ERROR',
                'sheet': sheet.title,
                'cell': row[3].coordinate,
                'message': 'Cell empty. Data mapping for this row is not possible.'
            })

        if row[5].value not in tier_list:
            error_list.append({
                'level': 'WARNING',
                'sheet': sheet.title,
                'cell': cell.coordinate,
                'message': f'Invalid Tier value. Please select a valid tier from the predefined tier list. (Current: "{cell.value}")'
            })

        for idx in range(9, 13):
            cell = row[idx]

            if cell.value is None:
                error_list.append({
                    'level': 'WARNING',
                    'sheet': sheet.title,
                    'cell': cell.coordinate,
                    'message': 'Cell empty. Value will be treated as 0.'
                })

            else:
                try:
                    float(cell.value)
                except (ValueError, TypeError):
                    error_list.append({
                        'level': 'CRITICAL',
                        'sheet': sheet.title,
                        'cell': cell.coordinate,
                        'message': f'Non-numeric value detected. This field only accepts numbers. (Current value: "{cell.value}")'
                    })

        for idx in range(len(header_list), len(header_list) + len(vertical_list)):
            cell = row[idx]

            if cell.value is not None and not isinstance(cell.value, bool):
                error_list.append({
                    'level': 'WARNING',
                    'sheet': sheet.title,
                    'cell': cell.coordinate,
                    'message': f'Invalid boolean value. "{cell.value}" will be treated as TRUE. Please enter TRUE or FALSE explicitly. (You can omit FALSE and leave cell empty for FALSE)'
                })

    return error_list