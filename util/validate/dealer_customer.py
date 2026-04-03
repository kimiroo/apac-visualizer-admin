from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from openpyxl.workbook.workbook import _WorksheetOrChartsheetLike

def check_dealer_customer_sheet(sheet: _WorksheetOrChartsheetLike, config: dict) -> list:
    error_list = [] # {'level': level, 'sheet': sheet, 'cell': cell, 'message': message}

    header_list = [
        'Dealer ID', 'Dealer Name', 'Name',
        'Sale Value', 'Sale Date', 'Sale Model'
    ]

    expected_max_col_cnt = len(header_list)

    ### Column count check
    if sheet.max_column != expected_max_col_cnt:
        error_list.append({
            'level': 'WARNING',
            'sheet': sheet.title,
            'cell': '-',
            'message': f'Column count mismatch. (Found: {sheet.max_column}, Expected: {expected_max_col_cnt})'
        })

    ### Header data check
    for idx, header in enumerate(header_list):
        col_idx = 1 + idx
        cell = sheet.cell(row=1, column=col_idx)

        if cell.value != header:
            error_list.append({
                'level': 'WARNING',
                'sheet': sheet.title,
                'cell': cell.coordinate,
                'message': f'Unexpected header value. Subsequent data may be invalid. (Found: {f'"{cell.value}"' if cell.value else 'None'}, Expected: "{header}")'
            })

    ### Data check
    for row in sheet.iter_rows(2):

        if row[0].value is None:
            error_list.append({
                'level': 'ERROR',
                'sheet': sheet.title,
                'cell': row[0].coordinate,
                'message': 'Cell empty. Data mapping for this row is not possible.'
            })

        if row[3].value is None:
            error_list.append({
                'level': 'WARNING',
                'sheet': sheet.title,
                'cell': row[3].coordinate,
                'message': 'Cell empty. Value will be treated as 0.'
            })

        else:
            try:
                float(row[3].value)
            except (ValueError, TypeError):
                error_list.append({
                    'level': 'CRITICAL',
                    'sheet': sheet.title,
                    'cell': row[3].coordinate,
                    'message': f'Non-numeric value detected. This field only accepts numbers. (Current value: "{row[3].value}")'
                })

        try:
            row[4].value.date()
        except AttributeError:
            error_list.append({
                'level': 'CRITICAL',
                'sheet': sheet.title,
                'cell': row[4].coordinate,
                'message': f'Invalid date format. This field only accepts valid Excel date values. (Current value: "{row[4].value}")'
            })

    return error_list