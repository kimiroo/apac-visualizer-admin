from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from openpyxl.workbook.workbook import _WorksheetOrChartsheetLike

def check_region_sheet(sheet: _WorksheetOrChartsheetLike, config: dict) -> list:
    error_list = [] # {'level': level, 'sheet': sheet, 'cell': cell, 'message': message}

    expected_max_col_cnt = 3 + ((len(config['vertical']) + 1) * 4)

    ### Column count check
    if sheet.max_column != expected_max_col_cnt:
        error_list.append({
            'level': 'WARNING',
            'sheet': sheet.title,
            'cell': '-',
            'message': f'Column count mismatch. (Found: {sheet.max_column}, Expected: {expected_max_col_cnt})'
        })

    ### Header data check
    for idx, header in enumerate(['Country', 'Dealer', 'Remarks']):
        col_idx = 1 + idx
        cell = sheet.cell(row=1, column=col_idx)

        if cell.value != header:
            error_list.append({
                'level': 'WARNING',
                'sheet': sheet.title,
                'cell': cell.coordinate,
                'message': f'Unexpected header value. Subsequent data may be invalid. (Found: {f'"{cell.value}"' if cell.value else 'None'}, Expected: "{header}")'
            })

    ### Vertical header data check
    headers = {}
    vertical_list = config['vertical'] + ['Total']

    col_start_idx = 3
    col_vertical = None

    # Parse merged header
    for idx, cell in enumerate(sheet[1]):
        if idx < 3:
            continue

        # New vertical column
        if ((idx - 3) % 4 == 0) and (idx != col_start_idx):
            # Append to headers
            headers[col_vertical] = col_start_idx

            # Reset variables
            col_start_idx = idx
            col_vertical = None

        # Skip to next column if valid vertical is already set
        if col_vertical:
            continue

        # New vertical found
        if cell.value in vertical_list:
            col_vertical = cell.value

    if col_vertical:
        # Append to headers
        headers[col_vertical] = col_start_idx

    for vertical in vertical_list:
        header_idx = headers.get(vertical, None)

        if header_idx is None:
            error_list.append({
                'level': 'CRITICAL',
                'sheet': sheet.title,
                'cell': '(Header Row)',
                'message': f'Required column "{vertical}" could not be found in the sheet headers.'
            })

    ### Data check
    for row in sheet.iter_rows(3):
        for idx in range(2):
            cell = row[idx]

            if cell.value is None:
                error_list.append({
                    'level': 'ERROR',
                    'sheet': sheet.title,
                    'cell': cell.coordinate,
                    'message': 'Cell empty. Data mapping for this row is not possible.'
                })

        vertical_col_cnt = (len(config['vertical']) + 1) * 4
        for idx in range(vertical_col_cnt):
            row_idx = 3 + idx
            cell = row[row_idx]

            if cell.value is None:
                error_list.append({
                    'level': 'WARNING',
                    'sheet': sheet.title,
                    'cell': cell.coordinate,
                    'message': 'Cell empty. Value will be treated as 0.'
                })

            else:
                try:
                    value = float(cell.value)

                    if (idx % 4) == 0 and not value.is_integer():
                        error_list.append({
                            'level': 'WARNING',
                            'sheet': sheet.title,
                            'cell': cell.coordinate,
                            'message': f'Decimal value detected in "Plant Count" field. Integer expected for this field. Any decimal values will be floored. (Current value: "{cell.value}")'
                        })

                except (ValueError, TypeError):
                    error_list.append({
                        'level': 'CRITICAL',
                        'sheet': sheet.title,
                        'cell': cell.coordinate,
                        'message': f'Non-numeric value detected. This field only accepts numbers. (Current value: "{cell.value}")'
                    })

    return error_list