from openpyxl import Workbook

from util.validate.region import check_region_sheet
from util.validate.dealer import check_dealer_sheet
from util.validate.dealer_customer import check_dealer_customer_sheet
from util.validate.key_account import check_key_account_sheet
from util.validate.priority_target import check_priority_target

def validate_data(doc:Workbook, config: dict) -> list:

    error_list = [] # {'level': level, 'sheet': sheet, 'cell': cell, 'message': message}

    ##############
    ### Region ###
    ##############

    sheet_name = config['source']['sheet']['region']['name']
    if not sheet_name in doc.sheetnames:
        error_list.append({
            'level': 'CRITICAL',
            'sheet': '-',
            'cell': '-',
            'message': f'Required sheet "{sheet_name}" not found in the workbook. Ensure the sheet name matches the configuration.'
        })
    else:
        sheet = doc[sheet_name]
        result_region = check_region_sheet(sheet, config)
        error_list = error_list + result_region

    ##############
    ### Dealer ###
    ##############

    sheet_name = config['source']['sheet']['dealer']['name']
    if not sheet_name in doc.sheetnames:
        error_list.append({
            'level': 'CRITICAL',
            'sheet': '-',
            'cell': '-',
            'message': f'Required sheet "{sheet_name}" not found in the workbook. Ensure the sheet name matches the configuration.'
        })
    else:
        sheet = doc[sheet_name]
        result_dealer = check_dealer_sheet(sheet, config)
        error_list = error_list + result_dealer

    #######################
    ### Dealer Customer ###
    #######################

    sheet_name = config['source']['sheet']['dealerCustomer']['name']
    if not sheet_name in doc.sheetnames:
        error_list.append({
            'level': 'CRITICAL',
            'sheet': '-',
            'cell': '-',
            'message': f'Required sheet "{sheet_name}" not found in the workbook. Ensure the sheet name matches the configuration.'
        })
    else:
        sheet = doc[sheet_name]
        result_dealer_customer = check_dealer_customer_sheet(sheet, config)
        error_list = error_list + result_dealer_customer

    ###################
    ### Key Account ###
    ###################

    sheet_name = config['source']['sheet']['keyAccount']['name']
    if not sheet_name in doc.sheetnames:
        error_list.append({
            'level': 'CRITICAL',
            'sheet': '-',
            'cell': '-',
            'message': f'Required sheet "{sheet_name}" not found in the workbook. Ensure the sheet name matches the configuration.'
        })
    else:
        sheet = doc[sheet_name]
        result_key_account = check_key_account_sheet(sheet, config)
        error_list = error_list + result_key_account

    ########################
    ### Priorioty Target ###
    ########################

    sheet_name = config['source']['sheet']['priorityTarget']['name']
    if not sheet_name in doc.sheetnames:
        error_list.append({
            'level': 'CRITICAL',
            'sheet': '-',
            'cell': '-',
            'message': f'Required sheet "{sheet_name}" not found in the workbook. Ensure the sheet name matches the configuration.'
        })
    else:
        sheet = doc[sheet_name]
        result_priority_target = check_priority_target(sheet, config)
        error_list = error_list + result_priority_target

    return error_list